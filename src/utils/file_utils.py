import os
import openpyxl as op
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import re 
import numpy as np 

FPS = 120
NUM_TRAILING_FILE_PARAMS_TO_IGNORE = 2

def get_all_videos(directory):
    return [video for video in os.listdir(directory) if is_video(video.lower())]

def is_video(video):
    is_video = False
    video_extensions = ["mov", "avi"]
    for video_extension in video_extensions:
        if video_extension in video:
            is_video = True 
            break

    return is_video

def open_prediction_file(file_path):
    '''
        Returns the contents of the prediction file.
    '''
    return np.load(file_path, allow_pickle=True).item()

def open_prediction_file_as_trial(file_path):
    '''
        Returns the contents of the prediction file and the length of the video. It also
        converts each prediction's frame number to its corresponding second.
    '''
    predictions = open_prediction_file(file_path)

    length_video = 0
    predictions_with_second_as_key = {}
    for frame_num, prediction_info in predictions.items():
        second = frame_num / FPS
        predictions_with_second_as_key.setdefault(second, prediction_info)

        if second > length_video:
            length_video = second

    return predictions_with_second_as_key, length_video

def open_prediction_file_as_trial_and_trim(file_path, desired_video_length):
    '''
        Returns the contents of the prediction file and the length of the video. It also
        converts each prediction's frame number to its corresponding second and trims
        the file to the desired video length.
    '''
    predictions = open_prediction_file(file_path)

    trimmed_predictions = {}
    length_video = 0
    desired_num_frames = desired_video_length * FPS

    if len(predictions) >= desired_num_frames:
        for frame_num, prediction_info in predictions.items():
            if frame_num <= desired_num_frames:
                second = frame_num / FPS
                trimmed_predictions.setdefault(second, prediction_info)
            else:
                length_video = second
                break

    return trimmed_predictions, length_video

def get_prediction_files(directory, one_file):
    if one_file:
        trial_files = get_all_files(directory)
        compass_files = trial_files
    else:
        trial_files = get_trial_files(directory)
        compass_files = get_compass_files(directory)

    return [trial_files, compass_files]

def get_compass_files(directory):
    return [file for file in get_all_files(directory) if is_compass_file(file)]

def get_trial_files(directory):
    return [file for file in get_all_files(directory) if not is_compass_file(file)]

def get_all_files(directory):
    try:
        files = [file for file in os.listdir(directory) if '.npy' in file]
    except FileNotFoundError:
        # write code to handle error
        files = []
    
    return files	

def is_compass_file(file):
    if 'compass' in file.lower():
        return True

    return False

def get_file_attributes(file):
    '''
        If present, it will return the date and time in the file's name.
    '''
    components = file.split('_')
    date = components[0]
    time = components[1]
    if date.isdigit() and time.isdigit():
        return int(date), int(time)
    else:
        return 19700101, 120000

def save_results(save_directory, all_trial_statistics, name_excel_file):
    wb = get_appropriate_workbook(all_trial_statistics)
    save_workbook(wb, save_directory, name_excel_file)

def get_appropriate_workbook(all_trial_statistics):
    if all_trial_statistics == {}:
        return create_default_workbook()
    else:
        return create_actual_workbook(all_trial_statistics)

def create_default_workbook():
    wb = op.Workbook()
    wb.create_sheet('data')
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    sheet = wb['data']
    sheet['A1'] = "Sorry, no statistics to output!"

    return wb 

def create_actual_workbook(all_trial_statistics):
    wb = op.Workbook()
    wb.create_sheet('data')
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    sheet = wb['data']

    headings = create_headings(all_trial_statistics)
    max_col_letter = get_column_letter(len(headings))

    write_headings_to_sheet(headings, sheet, max_col_letter)
    write_data_to_sheet(all_trial_statistics, sheet, max_col_letter)

    return wb 

def create_headings(all_trial_statistics):
    headings = ['file']

    sample_stat_set = all_trial_statistics[0][1]
    stat_names = [stat_name for stat_name in sample_stat_set.keys()]
    for stat_name in stat_names:
        headings.append(change_to_dot_notation(stat_name))

    files = [trial_statistics[0] for trial_statistics in all_trial_statistics]
    max_num_extra_headings = get_max_num_extra_headings(files)
    headings = add_empty_headings(headings, max_num_extra_headings)

    return headings

def change_to_dot_notation(stat_name):
    return stat_name.lower().replace("_", ".")

def get_max_num_extra_headings(files):
    num_extra_headings = [get_num_extra_headings(file) for file in files]
    return max(num_extra_headings)

def get_num_extra_headings(sample_file):
    return len(sample_file.split('_')[:-NUM_TRAILING_FILE_PARAMS_TO_IGNORE])

def add_empty_headings(headings, max_num_extra_headings):
    empty_headings = [''] * max_num_extra_headings
    return [headings[0]] + empty_headings + headings[1:]

def write_headings_to_sheet(headings, sheet, max_col_letter):
    col_num = 1
    for col in sheet['A1':f'{max_col_letter}1']:
        for cell in col:
            cell.value = headings[col_num - 1]
            col_num += 1

    return sheet 

def write_data_to_sheet(all_trial_statistics, sheet, max_col_letter):
    files = [trial_statistics[0] for trial_statistics in all_trial_statistics]
    stat_sets = [trial_statistics[1] for trial_statistics in all_trial_statistics]
    stat_names = [stat_name for stat_name in stat_sets[0].keys()]
    max_num_extra_headings = get_max_num_extra_headings(files)
    
    row_num = 0
    for row in sheet['A2':f'{max_col_letter}{2 + len(files) - 1}']:
        current_file = files[row_num]
        extra_headings = current_file.split('_')[:-NUM_TRAILING_FILE_PARAMS_TO_IGNORE]

        for cell in row:
            col_index = get_col_index(cell.coordinate)

            if col_index == 1:
                cell.value = current_file
            elif col_index <= max_num_extra_headings + 1:
                heading_index = col_index - 2
                try:
                    cell.value = extra_headings[heading_index]
                except IndexError:
                    pass
            else:
                stat_name_index = col_index - max_num_extra_headings - 2
                cell.value = stat_sets[row_num][stat_names[stat_name_index]]

        row_num += 1

    return sheet

def get_col_index(cell_coordinate):
    len_cell_coordinate = len(cell_coordinate)

    col_letter = ''
    for component in cell_coordinate:
        if component >= 'A' and component <= 'Z':
            col_letter += component

    return column_index_from_string(col_letter)


def save_workbook(wb, directory, name_excel_file):
    if directory != '':
        wb.save(f'{directory}/{name_excel_file}.xlsx')
    else:
        # write error to send to widget
        pass
